#!/usr/bin/env python3
"""Generate area-split BoQ CSVs (UTF-8 BOM) and a formatted Excel workbook."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path("/workspace/docs/procurement")

HEADERS = ["序号", "区域", "分部", "工作内容", "单位", "工程量", "单价", "合价", "备注"]

# row: (序号, 区域, 分部, 工作内容, 单位, 工程量, 单价, 合价, 备注)
# 工程量/单价/合价 empty unless header-type lump items that are clearly "1 item"

LAB = "25楼实验室"
OLD = "旧实验室修改"
WH = "一楼仓库"

BOQ_25F = [
    ("A", LAB, "开办费", "25楼实验室范围内的管理、垂直运输、通道及临时设施", "项", "1", "", "", "仅计本区域"),
    ("A.1", LAB, "开办费", "楼层成品保护、电梯及公共走道防护", "项", "1", "", "", ""),
    ("A.2", LAB, "开办费", "施工围挡 / 防尘隔断（与办公区交界）", "项", "1", "", "", ""),
    ("A.3", LAB, "开办费", "建筑垃圾分类清运（25楼）", "项", "1", "", "", ""),
    ("B", LAB, "隔墙及门窗", "实验室分区隔墙、观察窗、传递窗及实验室门", "项", "1", "", "", ""),
    ("B.1", LAB, "隔墙及门窗", "轻钢龙骨双面防火石膏板隔墙（含龙骨、填缝、阴角）", "m2", "", "", "", "按确认平面计量"),
    ("B.2", LAB, "隔墙及门窗", "实验室专用门（观察窗、闭门器、密封条）", "樘", "", "", "", ""),
    ("B.3", LAB, "隔墙及门窗", "防火门（前室/电气间/气瓶间，按消防审查）", "樘", "", "", "", "如适用"),
    ("B.4", LAB, "隔墙及门窗", "传递窗（机械或电子联锁）", "个", "", "", "", "如平面有"),
    ("B.5", LAB, "隔墙及门窗", "室内观察窗 / 固定玻璃隔断", "m2", "", "", "", ""),
    ("C", LAB, "地面", "实验室地面（含基层处理、踢脚）", "项", "1", "", "", ""),
    ("C.1", LAB, "地面", "原地面清理、找平、裂缝修补", "m2", "", "", "", ""),
    ("C.2", LAB, "地面", "实验室环氧树脂地坪（或等同等 PVC 卷材，待确认）", "m2", "", "", "", "请确认面层做法"),
    ("C.3", LAB, "地面", "防静电地坪（设备间，如需要）", "m2", "", "", "", "仅设备间"),
    ("C.4", LAB, "地面", "耐酸碱踢脚", "m", "", "", "", ""),
    ("C.5", LAB, "地面", "地漏及地面找坡（湿区）", "处", "", "", "", ""),
    ("D", LAB, "墙面及吊顶", "实验室墙面饰面及吊顶", "项", "1", "", "", ""),
    ("D.1", LAB, "墙面及吊顶", "耐酸碱墙面涂料 / 实验室墙板", "m2", "", "", "", "请确认涂料或墙板"),
    ("D.2", LAB, "墙面及吊顶", "铝扣板或洁净吊顶（含龙骨、检修口）", "m2", "", "", "", ""),
    ("D.3", LAB, "墙面及吊顶", "设备吊杆、管线穿吊顶封堵", "项", "1", "", "", ""),
    ("E", LAB, "实验室家具", "实验台、试剂柜、器皿柜及安装", "项", "1", "", "", "家具供货范围待确认"),
    ("E.1", LAB, "实验室家具", "边台 / 中央台（台面、支架、试剂架）", "m", "", "", "", "按布置图"),
    ("E.2", LAB, "实验室家具", "试剂柜 / 器皿柜 / 更衣柜", "个", "", "", "", ""),
    ("E.3", LAB, "实验室家具", "紧急冲淋 / 洗眼器安装配合（建筑开孔及固定）", "套", "", "", "", "设备如甲供请注明"),
    ("F", LAB, "通风柜及排风", "通风柜设备及实验室排风 / 补风", "项", "1", "", "", ""),
    ("F.1", LAB, "通风柜及排风", "通风柜供货及安装（含面风速调试）", "台", "", "", "", "数量按平面"),
    ("F.2", LAB, "通风柜及排风", "排风管道（含支架、防火阀）", "m", "", "", "", ""),
    ("F.3", LAB, "通风柜及排风", "补风 / 全新风管道", "m", "", "", "", ""),
    ("F.4", LAB, "通风柜及排风", "排风机 / 屋顶风机（含减振、风帽）", "台", "", "", "", "位置待确认"),
    ("F.5", LAB, "通风柜及排风", "房间压差及风量平衡调试", "项", "1", "", "", ""),
    ("G", LAB, "暖通空调", "实验室空调（舒适或工艺，待确认）", "项", "1", "", "", ""),
    ("G.1", LAB, "暖通空调", "室内机 / 末端 / 风口", "台", "", "", "", ""),
    ("G.2", LAB, "暖通空调", "冷媒管 / 冷凝水管", "m", "", "", "", ""),
    ("G.3", LAB, "暖通空调", "控制系统点位", "个", "", "", "", ""),
    ("H", LAB, "电气", "实验室强电及照明", "项", "1", "", "", ""),
    ("H.1", LAB, "电气", "实验室专用配电箱及断路器", "台", "", "", "", ""),
    ("H.2", LAB, "电气", "电缆、桥架及配管", "m", "", "", "", ""),
    ("H.3", LAB, "电气", "实验台下插座 / 墙上插座（含防水盖，按需）", "个", "", "", "", ""),
    ("H.4", LAB, "电气", "洁净灯盘 / 面板灯", "套", "", "", "", ""),
    ("H.5", LAB, "电气", "应急照明及出口指示", "套", "", "", "", ""),
    ("H.6", LAB, "电气", "等电位联结及接地", "项", "1", "", "", "实验室必需"),
    ("H.7", LAB, "电气", "电气测试与调试", "项", "1", "", "", ""),
    ("I", LAB, "给排水", "实验给水、纯水点位及废水", "项", "1", "", "", ""),
    ("I.1", LAB, "给排水", "实验台水槽、龙头及管道", "套", "", "", "", ""),
    ("I.2", LAB, "给排水", "纯水 / 去离子水点位配合", "个", "", "", "", "如系统甲供则仅配合"),
    ("I.3", LAB, "给排水", "耐酸碱废水管道", "m", "", "", "", ""),
    ("I.4", LAB, "给排水", "废水收集 / 中和装置（暂列）", "项", "1", "", "", "可选 — 待工艺确认"),
    ("I.5", LAB, "给排水", "紧急冲淋给水及排水", "套", "", "", "", ""),
    ("I.6", LAB, "给排水", "给排水试压及通水试验", "项", "1", "", "", ""),
    ("J", LAB, "特种气体", "气瓶间及气体管路（如适用）", "项", "1", "", "", "不用则整分部删除"),
    ("J.1", LAB, "特种气体", "气瓶固定及气体管路", "m", "", "", "", ""),
    ("J.2", LAB, "特种气体", "可燃 / 有毒气体报警探头", "个", "", "", "", ""),
    ("K", LAB, "弱电", "数据、门禁、监控及火灾报警", "项", "1", "", "", "请确认是否纳入本包"),
    ("K.1", LAB, "弱电", "数据点位", "个", "", "", "", ""),
    ("K.2", LAB, "弱电", "门禁点位", "个", "", "", "", ""),
    ("K.3", LAB, "弱电", "监控点位", "个", "", "", "", ""),
    ("K.4", LAB, "弱电", "火灾报警探测器 / 手报 / 声光", "个", "", "", "", ""),
    ("L", LAB, "消防及其他", "防火封堵、灭火器及交工清洁", "项", "1", "", "", ""),
    ("L.1", LAB, "消防及其他", "套管及孔洞防火封堵", "项", "1", "", "", ""),
    ("L.2", LAB, "消防及其他", "灭火器箱及配置", "套", "", "", "", ""),
    ("L.3", LAB, "消防及其他", "竣工清洁及成品保护拆除", "项", "1", "", "", ""),
    ("M", LAB, "暂列", "25楼实验室不可预见费（可选）", "笔", "1", "", "", "不需要则删除"),
]

BOQ_OLD = [
    ("A", OLD, "保护及隔离", "旧实验室施工期间对保留设备、管线及相邻房间的保护", "项", "1", "", "", ""),
    ("A.1", OLD, "保护及隔离", "现有仪器、台柜覆膜及硬质围挡", "项", "1", "", "", "与使用中区域交界必须做"),
    ("A.2", OLD, "保护及隔离", "施工隔离、防尘及临时通道", "项", "1", "", "", ""),
    ("B", OLD, "拆除工程", "按改造范围拆除隔墙、吊顶、地面、台柜及废弃管线", "项", "1", "", "", "拆除前请确认保留清单"),
    ("B.1", OLD, "拆除工程", "轻质隔墙拆除（含运出）", "m2", "", "", "", ""),
    ("B.2", OLD, "拆除工程", "吊顶局部拆除", "m2", "", "", "", ""),
    ("B.3", OLD, "拆除工程", "地面面层局部铲除", "m2", "", "", "", ""),
    ("B.4", OLD, "拆除工程", "实验台柜拆除或拆卸移位（不含新购）", "项", "1", "", "", "请确认移位还是报废"),
    ("B.5", OLD, "拆除工程", "废弃电气点位拆除及封堵", "个", "", "", "", ""),
    ("B.6", OLD, "拆除工程", "废弃给排水 / 排风管拆除", "m", "", "", "", ""),
    ("B.7", OLD, "拆除工程", "建筑垃圾清运", "项", "1", "", "", ""),
    ("C", OLD, "建筑改造", "隔墙、门窗、地面、墙面、吊顶按改造后平面恢复或新做", "项", "1", "", "", ""),
    ("C.1", OLD, "建筑改造", "补做 / 改做轻钢龙骨隔墙", "m2", "", "", "", ""),
    ("C.2", OLD, "建筑改造", "门扇改位、更换或封堵门洞", "樘", "", "", "", ""),
    ("C.3", OLD, "建筑改造", "地面修补及局部环氧 / PVC 接茬", "m2", "", "", "", "与保留地面颜色衔接"),
    ("C.4", OLD, "建筑改造", "墙面修补及耐酸碱涂料重涂", "m2", "", "", "", ""),
    ("C.5", OLD, "建筑改造", "吊顶修补、调平及更换破损板块", "m2", "", "", "", ""),
    ("D", OLD, "台柜改造", "保留实验台柜的改位、拼接及台面修补", "项", "1", "", "", ""),
    ("D.1", OLD, "台柜改造", "实验台移位、重新固定及调平", "m", "", "", "", ""),
    ("D.2", OLD, "台柜改造", "台面切割、封边及水槽孔改制", "项", "1", "", "", ""),
    ("D.3", OLD, "台柜改造", "新增短段边台（仅改造缺口补齐）", "m", "", "", "", "大面积新做计入25楼单"),
    ("E", OLD, "电气改造", "回路改位、配电箱调整及照明点位修改", "项", "1", "", "", ""),
    ("E.1", OLD, "电气改造", "插座 / 回路改位或新增", "个", "", "", "", ""),
    ("E.2", OLD, "电气改造", "灯位改位或更换", "套", "", "", "", ""),
    ("E.3", OLD, "电气改造", "现有配电箱回路调整", "项", "1", "", "", "以不换箱为原则，除非容量不够"),
    ("E.4", OLD, "电气改造", "改完后绝缘及通电测试", "项", "1", "", "", ""),
    ("F", OLD, "暖通排风改造", "风管改位、通风柜移位及风量重新平衡", "项", "1", "", "", ""),
    ("F.1", OLD, "暖通排风改造", "通风柜移位（含重新连接排风）", "台", "", "", "", "如柜体报废则删此行"),
    ("F.2", OLD, "暖通排风改造", "排风 / 补风管道改位", "m", "", "", "", ""),
    ("F.3", OLD, "暖通排风改造", "空调室内机改位或风口调整", "台", "", "", "", ""),
    ("F.4", OLD, "暖通排风改造", "风量、压差重新调试", "项", "1", "", "", ""),
    ("G", OLD, "给排水改造", "水槽、龙头及废水管改位", "项", "1", "", "", ""),
    ("G.1", OLD, "给排水改造", "冷热水点改位", "个", "", "", "", ""),
    ("G.2", OLD, "给排水改造", "废水管道改管（耐酸碱）", "m", "", "", "", ""),
    ("G.3", OLD, "给排水改造", "地漏改位及地面找坡修补", "处", "", "", "", ""),
    ("G.4", OLD, "给排水改造", "改完后试压、通水", "项", "1", "", "", ""),
    ("H", OLD, "弱电改造", "数据、门禁、监控、火警点位随隔墙改位", "项", "1", "", "", "请确认是否纳入本包"),
    ("H.1", OLD, "弱电改造", "数据点改位 / 恢复", "个", "", "", "", ""),
    ("H.2", OLD, "弱电改造", "门禁 / 监控改位", "个", "", "", "", ""),
    ("H.3", OLD, "弱电改造", "火灾报警探测器改位及复位", "个", "", "", "", ""),
    ("I", OLD, "修复及清洁", "与保留区域接茬、洞口封堵、清洁交工", "项", "1", "", "", ""),
    ("I.1", OLD, "修复及清洁", "孔洞防火封堵及装饰收口", "项", "1", "", "", ""),
    ("I.2", OLD, "修复及清洁", "与未改区域的饰面、地面接茬处理", "项", "1", "", "", "避免新旧明显色差"),
    ("I.3", OLD, "修复及清洁", "竣工清洁、保护拆除、设备复位配合", "项", "1", "", "", ""),
    ("J", OLD, "暂列", "旧实验室修改不可预见费（可选，隐蔽管线风险）", "笔", "1", "", "", "不需要则删除"),
]

BOQ_WH = [
    ("A", WH, "开办费", "一楼仓库范围内的管理、卸货通道、围挡及临时设施", "项", "1", "", "", "仅计本区域"),
    ("A.1", WH, "开办费", "一楼卸货口占用、临时围挡及交通引导", "项", "1", "", "", ""),
    ("A.2", WH, "开办费", "成品保护（柱面、门口、已完地坪）", "项", "1", "", "", ""),
    ("A.3", WH, "开办费", "建筑垃圾清运（一楼）", "项", "1", "", "", ""),
    ("B", WH, "地面", "仓库地坪、找平、面层及划线", "项", "1", "", "", ""),
    ("B.1", WH, "地面", "原地面清理、空鼓修补、找平", "m2", "", "", "", ""),
    ("B.2", WH, "地面", "混凝土补强 / 硬化剂地坪（如需要）", "m2", "", "", "", "请确认是否已有结构地坪"),
    ("B.3", WH, "地面", "环氧地坪（或耐磨地坪，待确认）", "m2", "", "", "", ""),
    ("B.4", WH, "地面", "车位 / 货位 / 通道划线及编号", "m", "", "", "", ""),
    ("B.5", WH, "地面", "防撞角钢 / 柱角护角", "处", "", "", "", ""),
    ("C", WH, "墙面及门", "墙面涂料、防撞及大门", "项", "1", "", "", ""),
    ("C.1", WH, "墙面及门", "内墙抹灰修补及涂料（含防火涂料，按需）", "m2", "", "", "", ""),
    ("C.2", WH, "墙面及门", "电动卷帘门（含导轨、电机、控制）", "樘", "", "", "", "卸货口"),
    ("C.3", WH, "墙面及门", "人行防火门 / 疏散门", "樘", "", "", "", ""),
    ("C.4", WH, "墙面及门", "门斗、雨篷或卸货平台防雨配合", "项", "1", "", "", "如现场有"),
    ("D", WH, "货架配合", "货架基础、预埋及安装配合（不含货架供货，除非确认纳入）", "项", "1", "", "", "货架供货请采购确认"),
    ("D.1", WH, "货架配合", "货架地脚螺栓预埋 / 植筋", "处", "", "", "", ""),
    ("D.2", WH, "货架配合", "局部基础加厚或垫层", "m3", "", "", "", "按货架厂家荷载"),
    ("D.3", WH, "货架配合", "货架供货及安装（暂列，可选）", "项", "1", "", "", "不在本合同则删除"),
    ("E", WH, "电气照明", "仓库动力配电及高天井照明", "项", "1", "", "", ""),
    ("E.1", WH, "电气照明", "仓库配电箱", "台", "", "", "", ""),
    ("E.2", WH, "电气照明", "电缆、桥架及配管", "m", "", "", "", ""),
    ("E.3", WH, "电气照明", "高天井灯 / LED 工矿灯", "套", "", "", "", ""),
    ("E.4", WH, "电气照明", "应急照明及出口指示", "套", "", "", "", ""),
    ("E.5", WH, "电气照明", "叉车 / 充电或动力插座（如需要）", "个", "", "", "", ""),
    ("E.6", WH, "电气照明", "电气测试与调试", "项", "1", "", "", ""),
    ("F", WH, "通风", "仓库通风（换气扇或风管，按层高及存货）", "项", "1", "", "", "如仅自然通风可删"),
    ("F.1", WH, "通风", "排风扇 / 屋顶风机", "台", "", "", "", ""),
    ("F.2", WH, "通风", "风管及风口（如有）", "m", "", "", "", ""),
    ("G", WH, "消防", "仓库消防喷淋、感烟、灭火器及防火封堵", "项", "1", "", "", "按消防审查"),
    ("G.1", WH, "消防", "喷淋管道、喷头及支架", "m", "", "", "", "如原系统可接驳请注明"),
    ("G.2", WH, "消防", "感烟 / 感温探测器", "个", "", "", "", ""),
    ("G.3", WH, "消防", "消火栓箱改位或新增（如需要）", "套", "", "", "", ""),
    ("G.4", WH, "消防", "灭火器箱及配置", "套", "", "", "", ""),
    ("G.5", WH, "消防", "防火封堵", "项", "1", "", "", ""),
    ("H", WH, "弱电", "监控、门禁及数据（仓库办公角，如有）", "项", "1", "", "", "请确认是否纳入本包"),
    ("H.1", WH, "弱电", "监控点位", "个", "", "", "", "卸货口及货区"),
    ("H.2", WH, "弱电", "门禁点位", "个", "", "", "", ""),
    ("H.3", WH, "弱电", "数据点位（值班/办公小间）", "个", "", "", "", ""),
    ("I", WH, "辅房（可选）", "仓库内值班室 / 办公小间隔墙及简单装修", "项", "1", "", "", "没有辅房则整分部删除"),
    ("I.1", WH, "辅房（可选）", "轻质隔墙及门", "m2", "", "", "", ""),
    ("I.2", WH, "辅房（可选）", "简易吊顶、地面、灯具", "m2", "", "", "", ""),
    ("J", WH, "标识及交工", "标识标牌、竣工清洁", "项", "1", "", "", ""),
    ("J.1", WH, "标识及交工", "安全标识、限高、禁放、疏散图", "项", "1", "", "", ""),
    ("J.2", WH, "标识及交工", "竣工清洁及保护拆除", "项", "1", "", "", ""),
    ("K", WH, "暂列", "一楼仓库不可预见费（可选）", "笔", "1", "", "", "不需要则删除"),
]

SUMMARY = [
    ("A", LAB, "本区域小计", "25楼实验室全部补充工作（见分项表）", "项", "1", "", "", "金额引自「25楼实验室」表"),
    ("B", OLD, "本区域小计", "旧实验室修改全部补充工作（见分项表）", "项", "1", "", "", "金额引自「旧实验室修改」表"),
    ("C", WH, "本区域小计", "一楼仓库全部补充工作（见分项表）", "项", "1", "", "", "金额引自「一楼仓库」表"),
    ("D", "三份合计", "合计", "上述 A+B+C，仅为额外/补充范围", "项", "1", "", "", "不含原合同金额"),
]

DWG_25F = [
    ("A-L25-001", LAB, "25楼实验室 — 平面布置图（草稿）", "建筑", "A", "2026-09-03", "草稿 — 待审核", "drawings/25f-lab/", "隔墙、功能分区、门窗"),
    ("A-L25-002", LAB, "25楼实验室 — 实验台及家具布置图", "建筑", "A", "2026-09-03", "草稿 / 待定", "drawings/25f-lab/", ""),
    ("A-L25-003", LAB, "25楼实验室 — 天花 / 灯具平面图", "建筑", "A", "2026-09-03", "未开始 / 待定", "drawings/25f-lab/", ""),
    ("E-L25-001", LAB, "25楼实验室 — 电气平面图", "机电", "A", "2026-09-03", "草稿 / 待定", "drawings/25f-lab/", "插座、照明、配电"),
    ("M-L25-001", LAB, "25楼实验室 — 排风 / 暖通平面图", "机电", "A", "2026-09-03", "草稿 / 待定", "drawings/25f-lab/", "通风柜及排风"),
    ("P-L25-001", LAB, "25楼实验室 — 给排水平面图", "机电", "A", "2026-09-03", "草稿 / 待定", "drawings/25f-lab/", "实验水槽、废水、冲淋"),
]
DWG_OLD = [
    ("A-OL-001", OLD, "旧实验室 — 现状 / 拆除范围图", "建筑", "A", "2026-09-03", "草稿 — 待审核", "drawings/old-lab/", "标明保留 / 拆除"),
    ("A-OL-002", OLD, "旧实验室 — 改造后平面图", "建筑", "A", "2026-09-03", "草稿 — 待审核", "drawings/old-lab/", "与拆除图对照"),
    ("A-OL-003", OLD, "旧实验室 — 台柜及管线改位对照", "草图", "A", "2026-09-03", "草稿 / 待定", "drawings/old-lab/", "避免与25楼新做重复计量"),
    ("E-OL-001", OLD, "旧实验室 — 电气改位平面图", "机电", "A", "2026-09-03", "草稿 / 待定", "drawings/old-lab/", ""),
    ("M-OL-001", OLD, "旧实验室 — 排风 / 空调改位平面图", "机电", "A", "2026-09-03", "草稿 / 待定", "drawings/old-lab/", ""),
]
DWG_WH = [
    ("A-WH-001", WH, "一楼仓库 — 平面布置图（草稿）", "建筑", "A", "2026-09-03", "草稿 — 待审核", "drawings/1f-warehouse/", "大门、货区、辅房"),
    ("A-WH-002", WH, "一楼仓库 — 地坪划线及货架布置", "建筑", "A", "2026-09-03", "草稿 / 待定", "drawings/1f-warehouse/", "货架若甲供仍需此图配合"),
    ("E-WH-001", WH, "一楼仓库 — 电气照明平面图", "机电", "A", "2026-09-03", "草稿 / 待定", "drawings/1f-warehouse/", "高天井灯、动力"),
    ("F-WH-001", WH, "一楼仓库 — 消防平面图", "机电", "A", "2026-09-03", "草稿 / 待定", "drawings/1f-warehouse/", "喷淋、感烟、疏散"),
]
DWG_SHARED = [
    ("SK-ALL-001", "三区交界", "范围边界 / 与原合同交界面草图", "草图", "A", "2026-09-03", "草稿 — 待审核", "drawings/", "防止三份之间或与原合同重复计量"),
]
DRAWINGS = DWG_25F + DWG_OLD + DWG_WH + DWG_SHARED

DRAWING_HEADERS = ["图号", "区域", "图名", "类型", "版本", "日期", "状态", "文件/位置", "备注"]


def write_csv(path: Path, headers: list[str], rows: list[tuple]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


AREA_FILLS = {
    LAB: PatternFill("solid", fgColor="D6EAF8"),
    OLD: PatternFill("solid", fgColor="FCF3CF"),
    WH: PatternFill("solid", fgColor="D5F5E3"),
    "三份合计": PatternFill("solid", fgColor="FADBD8"),
    "三区交界": PatternFill("solid", fgColor="E8DAEF"),
}

HEADER_ROW_FILL = PatternFill("solid", fgColor="F2F3F4")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
FONT = Font(name="Microsoft YaHei", size=10)
FONT_BOLD = Font(name="Microsoft YaHei", size=10, bold=True)
WRAP = Alignment(vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def is_section_row(seq: str) -> bool:
    return seq and "." not in str(seq)


def write_boq_sheet(ws, title: str, rows: list[tuple]) -> int:
    """Write BoQ rows starting at row 1. Returns the totals row index."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="1F4E79")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "额外 / 补充范围 · 工作草稿 · 工程量待实量 · 单价建议与原合同同类工种一致"
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")
    ws.merge_cells("A2:I2")

    header_row = 4
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(header_row, c, h)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = "A5"

    for i, row in enumerate(rows):
        r = header_row + 1 + i
        seq, area, trade, desc, unit, qty, rate, amount, note = row
        values = [seq, area, trade, desc, unit, qty, rate, amount, note]
        section = is_section_row(seq)
        fill = AREA_FILLS.get(area)
        for c, val in enumerate(values, 1):
            cell = ws.cell(r, c, val if val != "" else None)
            cell.font = FONT_BOLD if section else FONT
            cell.alignment = CENTER if c in (1, 2, 5) else WRAP
            cell.border = THIN
            if section:
                cell.fill = HEADER_ROW_FILL
            elif fill and c == 2:
                cell.fill = fill
            if c in (6, 7, 8) and val not in ("", None):
                try:
                    cell.value = float(val)
                except ValueError:
                    pass
            if c in (6, 7, 8):
                cell.number_format = "#,##0.00"
                cell.alignment = CENTER
        # 合价公式
        ws.cell(r, 8).value = f"=IF(OR(F{r}=\"\",G{r}=\"\"),\"\",F{r}*G{r})"
        ws.cell(r, 8).number_format = "#,##0.00"
        ws.row_dimensions[r].height = 32 if section else 36

    last = header_row + len(rows)
    total_row = last + 1
    for c in range(1, 10):
        ws.cell(total_row, c).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(total_row, c).font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        ws.cell(total_row, c).border = THIN
    ws.cell(total_row, 1, f"{title} 合价小计")
    ws.cell(total_row, 1).alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    ws.cell(total_row, 8, f"=SUM(H{header_row+1}:H{last})")
    ws.cell(total_row, 8).number_format = '#,##0.00'
    ws.cell(total_row, 8).font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    ws.cell(total_row, 9, "未填单价时为 0，不代表免费")
    ws.cell(total_row, 9).font = Font(name="Microsoft YaHei", size=9, color="FFFFFF")
    ws.row_dimensions[total_row].height = 22

    widths = {"A": 10, "B": 16, "C": 16, "D": 52, "E": 8, "F": 10, "G": 12, "H": 14, "I": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A{header_row}:I{last}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddHeader.left.text = "补充报价 · 仅供内部确认"
    ws.oddFooter.center.text = "&P / &N"
    return total_row


def write_summary_sheet(ws, total_cells: dict[str, str]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "补充报价与图纸总览 — 已按区域拆成三份"
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="1F4E79")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "日期：2026年9月3日    目的：采购内部确认额外工程范围    报价与图纸按同一区域拆分    不含原合同金额"
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")
    ws.merge_cells("A2:G2")

    notes = [
        "A  25楼实验室：报价 + 图纸（平面、台柜、排风、实验水电）。独立文件：A-25楼实验室-报价与图纸.xlsx",
        "B  旧实验室修改：报价 + 图纸（拆除范围、改造后平面、管线改位）。独立文件：B-旧实验室修改-报价与图纸.xlsx",
        "C  一楼仓库：报价 + 图纸（平面、划线货架、照明、消防）。独立文件：C-一楼仓库-报价与图纸.xlsx",
    ]
    for i, text in enumerate(notes):
        ws.cell(4 + i, 1, text)
        ws.merge_cells(start_row=4 + i, start_column=1, end_row=4 + i, end_column=7)
        ws.cell(4 + i, 1).font = FONT
        ws.cell(4 + i, 1).alignment = WRAP
        ws.row_dimensions[4 + i].height = 20

    headers = ["编号", "区域", "工作表", "范围摘要", "合价", "状态（采购填写）", "备注"]
    hr = 8
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A9"

    data = [
        ("A", LAB, "25楼实验室", "新实验室建筑装饰、台柜、通风柜、实验水电", total_cells[LAB], "待确认", "见工作表「25楼实验室」"),
        ("B", OLD, "旧实验室修改", "保护、拆除、改位、修复", total_cells[OLD], "待确认", "见工作表「旧实验室修改」"),
        ("C", WH, "一楼仓库", "地坪、大门、照明、动力、消防", total_cells[WH], "待确认", "见工作表「一楼仓库」"),
    ]
    for i, (code, area, sheet, summary, formula, status, note) in enumerate(data):
        r = hr + 1 + i
        vals = [code, area, sheet, summary, None, status, note]
        fill = AREA_FILLS[area]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val)
            cell.font = FONT_BOLD if c <= 2 else FONT
            cell.alignment = CENTER if c in (1, 2, 3, 6) else WRAP
            cell.border = THIN
            cell.fill = fill
        ws.cell(r, 5, formula)
        ws.cell(r, 5).number_format = '#,##0.00'
        ws.cell(r, 5).font = FONT_BOLD
        ws.row_dimensions[r].height = 28

    total_r = hr + 4
    ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=4)
    ws.cell(total_r, 1, "三份合计（额外 / 补充范围）")
    ws.cell(total_r, 1).font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    ws.cell(total_r, 1).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(1, 8):
        ws.cell(total_r, c).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(total_r, c).border = THIN
        ws.cell(total_r, c).font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    ws.cell(total_r, 5, f"=SUM(E{hr+1}:E{hr+3})")
    ws.cell(total_r, 5).number_format = '#,##0.00'
    ws.cell(total_r, 6, "待确认")
    ws.cell(total_r, 7, "未填单价时合计为 0，不代表报价为零")

    dv = DataValidation(type="list", formula1='"待确认,确认,修改,不需要"', allow_blank=True)
    dv.error = "请选择 待确认 / 确认 / 修改 / 不需要"
    dv.errorTitle = "状态"
    dv.prompt = "采购填写"
    dv.promptTitle = "本份意见"
    ws.add_data_validation(dv)
    dv.add(f"F{hr+1}:F{hr+3}")

    ws["A14"] = "填写说明"
    ws["A14"].font = Font(name="Microsoft YaHei", size=12, bold=True, color="1F4E79")
    steps = [
        "1. 打开对应区域工作表，按实量填写「工程量」；采购要求带价时再填「单价」，「合价」自动计算。",
        "2. 不属于该区域的行请删除（例如仓库单里不要留通风柜；25楼单里不要留旧实验室拆除）。",
        "3. 本页「状态」列请采购按份选择：确认 / 修改 / 不需要。",
        "4. 图纸与报价同一区域：25楼放 drawings/25f-lab，旧实验室放 drawings/old-lab，仓库放 drawings/1f-warehouse。",
        "5. 发给采购时优先附三份独立文件（A/B/C 各含报价单+图纸目录），不要再发混在一起的旧版。",
        "6. 货架供货、实验废水中和、特种气体、辅房隔间等为可选项，不需要则删除对应行。",
    ]
    for i, s in enumerate(steps):
        ws.cell(15 + i, 1, s)
        ws.merge_cells(start_row=15 + i, start_column=1, end_row=15 + i, end_column=7)
        ws.cell(15 + i, 1).font = FONT
        ws.row_dimensions[15 + i].height = 20

    widths = {"A": 10, "B": 16, "C": 16, "D": 44, "E": 14, "F": 18, "G": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def write_drawing_sheet(ws, title: str, rows: list[tuple]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="1F4E79")
    ws.merge_cells("A1:I1")
    ws["A2"] = "尚未完成的图不要附在邮件里，目录中标注为「未开始 / 待定」即可。图纸与报价同一区域，不要混放。"
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")
    ws.merge_cells("A2:I2")

    hr = 4
    for c, h in enumerate(DRAWING_HEADERS, 1):
        cell = ws.cell(hr, c, h)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A5"
    for i, row in enumerate(rows):
        r = hr + 1 + i
        area = row[1]
        fill = AREA_FILLS.get(area)
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = FONT
            cell.alignment = CENTER if c in (1, 2, 4, 5, 6) else WRAP
            cell.border = THIN
            if fill and c == 2:
                cell.fill = fill
        ws.row_dimensions[r].height = 32
    last = hr + len(rows)
    widths = {"A": 14, "B": 16, "C": 42, "D": 10, "E": 8, "F": 14, "G": 16, "H": 22, "I": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A{hr}:I{last}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def write_area_pack(path: Path, title: str, boq_rows: list[tuple], dwg_rows: list[tuple]) -> None:
    """One area = one file: quote sheet + drawing register."""
    wb = Workbook()
    ws_boq = wb.active
    ws_boq.title = "报价单"
    write_boq_sheet(ws_boq, title, boq_rows)
    ws_dwg = wb.create_sheet("图纸目录")
    write_drawing_sheet(ws_dwg, f"{title} — 图纸目录", dwg_rows)
    wb.save(path)
    print(f"wrote {path}")


def main() -> None:
    write_csv(ROOT / "02-boq-summary.csv", HEADERS, SUMMARY)
    write_csv(ROOT / "02a-boq-25f-lab.csv", HEADERS, BOQ_25F)
    write_csv(ROOT / "02b-boq-old-lab.csv", HEADERS, BOQ_OLD)
    write_csv(ROOT / "02c-boq-1f-warehouse.csv", HEADERS, BOQ_WH)
    write_csv(ROOT / "03a-drawings-25f-lab.csv", DRAWING_HEADERS, DWG_25F)
    write_csv(ROOT / "03b-drawings-old-lab.csv", DRAWING_HEADERS, DWG_OLD)
    write_csv(ROOT / "03c-drawings-1f-warehouse.csv", DRAWING_HEADERS, DWG_WH)
    write_csv(ROOT / "03-drawing-register.csv", DRAWING_HEADERS, DRAWINGS)

    write_area_pack(ROOT / "A-25楼实验室-报价与图纸.xlsx", LAB, BOQ_25F, DWG_25F)
    write_area_pack(ROOT / "B-旧实验室修改-报价与图纸.xlsx", OLD, BOQ_OLD, DWG_OLD)
    write_area_pack(ROOT / "C-一楼仓库-报价与图纸.xlsx", WH, BOQ_WH, DWG_WH)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "总览"

    ws_lab = wb.create_sheet("25楼实验室")
    ws_old = wb.create_sheet("旧实验室修改")
    ws_wh = wb.create_sheet("一楼仓库")
    ws_dwg_lab = wb.create_sheet("25楼图纸")
    ws_dwg_old = wb.create_sheet("旧实验室图纸")
    ws_dwg_wh = wb.create_sheet("一楼仓库图纸")
    ws_dwg_all = wb.create_sheet("图纸总目录")

    lab_total_row = write_boq_sheet(ws_lab, "25楼实验室", BOQ_25F)
    old_total_row = write_boq_sheet(ws_old, "旧实验室修改", BOQ_OLD)
    wh_total_row = write_boq_sheet(ws_wh, "一楼仓库", BOQ_WH)

    total_cells = {
        LAB: f"='25楼实验室'!H{lab_total_row}",
        OLD: f"='旧实验室修改'!H{old_total_row}",
        WH: f"='一楼仓库'!H{wh_total_row}",
    }
    write_summary_sheet(ws_sum, total_cells)
    write_drawing_sheet(ws_dwg_lab, "25楼实验室 — 图纸目录", DWG_25F)
    write_drawing_sheet(ws_dwg_old, "旧实验室修改 — 图纸目录", DWG_OLD)
    write_drawing_sheet(ws_dwg_wh, "一楼仓库 — 图纸目录", DWG_WH)
    write_drawing_sheet(ws_dwg_all, "图纸总目录 — 三区 + 交界面", DRAWINGS)

    out = ROOT / "报价与图纸-按区域拆分.xlsx"
    wb.save(out)
    print(f"wrote {out}")
    print(f"25F rows={len(BOQ_25F)} old={len(BOQ_OLD)} wh={len(BOQ_WH)} dwg={len(DRAWINGS)}")


if __name__ == "__main__":
    main()
